"""交付物评测集（M3-7 / R12）：四种格式各一条 e2e——真实数据生成 + 程序化回读校验。

用法：DATABASE_URL=... .venv/bin/python eval/deliverable_eval.py
判定：文件产出在工作区 work/ + 可被对应库打开 + 结构/内容达标。
"""

import asyncio
import sys
import time
from pathlib import Path

sys.path.insert(0, "src")

from langgraph.checkpoint.memory import InMemorySaver
from langgraph.store.memory import InMemoryStore

from self_agent import settings
from self_agent.agent import build_agent
from self_agent.mcp import load_mcp_tools
from self_agent.project import load_project

WORK = settings.WORKSPACE_ROOT / "uav" / "work"  # 项目化后工作区在项目子目录


def _verify_xlsx(p: Path) -> str | None:
    from openpyxl import load_workbook

    wb = load_workbook(p)
    ws = wb[wb.sheetnames[0]]
    if ws.max_row < 2:
        return f"数据行不足: {ws.max_row}"
    return None


def _verify_pptx(p: Path) -> str | None:
    from pptx import Presentation

    n = len(Presentation(p).slides._sldIdLst)
    return None if n >= 2 else f"页数不足: {n}"


def _verify_docx(p: Path) -> str | None:
    from docx import Document

    doc = Document(p)
    text = "\n".join(x.text for x in doc.paragraphs)
    return None if len(text) > 50 else "正文过短"


def _verify_pdf(p: Path) -> str | None:
    head = p.read_bytes()[:5]
    if head != b"%PDF-":
        return "非 PDF 文件头"
    return None if p.stat().st_size > 3000 else "文件过小"


CASES = [
    ("xlsx", "查汉川市的图斑，把清单整理成 Excel 表格", _verify_xlsx),
    ("pptx", "查一下当前未处理告警的概况，出一份两页左右的简报 PPT", _verify_pptx),
    ("docx", "查庙头镇机场的状态，写一份简短的设备核查记录 Word 文档", _verify_docx),
    ("pdf", "查汉川市的禁飞区情况，整理成一份 PDF 简报", _verify_pdf),
]


async def main():
    PROJECT = load_project("uav")
    tools, down = await load_mcp_tools(PROJECT)
    agent = build_agent(PROJECT, tools, down_domains=down, checkpointer=InMemorySaver(),
                        store=InMemoryStore())
    WORK.mkdir(parents=True, exist_ok=True)
    passed = 0
    for ext, prompt, verify in CASES:
        before = {f: f.stat().st_mtime for f in WORK.glob(f"*.{ext}")}
        t0 = time.time()
        try:
            await asyncio.wait_for(agent.ainvoke(
                {"messages": [("user", prompt)]},
                config={"configurable": {"thread_id": f"deliv-{ext}"}, "recursion_limit": 160}  # 与 ModelCallLimit(80) 配套，评测不先于框架限额卡住,
            ), timeout=420)
        except Exception as e:  # noqa: BLE001
            print(f"❌ {ext}: run 异常 {type(e).__name__}: {str(e)[:80]}")
            continue
        new = [f for f in WORK.glob(f"*.{ext}")
               if f not in before or f.stat().st_mtime > before[f]]
        if not new:
            print(f"❌ {ext}: 未产出文件（{time.time()-t0:.0f}s）")
            continue
        target = max(new, key=lambda f: f.stat().st_mtime)
        err = verify(target)
        if err:
            print(f"❌ {ext}: {target.name} 校验失败——{err}")
        else:
            passed += 1
            print(f"✅ {ext}: {target.name}（{target.stat().st_size//1024}KB，{time.time()-t0:.0f}s）")
    print(f"\n== 交付物评测: {passed}/{len(CASES)} ==")


if __name__ == "__main__":
    asyncio.run(main())
