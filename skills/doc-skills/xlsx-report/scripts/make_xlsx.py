"""JSON 规格 → xlsx。用法：python make_xlsx.py <spec.json> <out.xlsx>"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def main():
    spec_path, out_path = sys.argv[1], sys.argv[2]
    spec = json.loads(Path(spec_path).read_text())
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in spec.get("sheets", []):
        ws = wb.create_sheet(str(sheet.get("name", "Sheet"))[:31])
        headers = sheet.get("headers") or []
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=str(h))
            cell.fill = header_fill
            cell.font = header_font
        widths = [len(str(h)) * 2 for h in headers]
        for r, row in enumerate(sheet.get("rows") or [], start=2):
            for c, val in enumerate(row[: len(headers) or None], start=1):
                ws.cell(row=r, column=c, value=val)
                if c <= len(widths):
                    widths[c - 1] = max(widths[c - 1], len(str(val)) + 2)
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = min(w, 50)
        ws.freeze_panes = "A2"
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"OK {out_path} {len(wb.sheetnames)}表")


if __name__ == "__main__":
    main()
